"""
Import Chinese rank names from fenlei.xlsx into taxon_name_zh table.

Usage:
  # Import directly into running DB (idempotent UPSERT):
  docker compose exec backend python -m scripts.import_taxon_zh

  # Regenerate the static seed SQL file (for fresh deployments):
  docker compose exec backend python -m scripts.import_taxon_zh --emit-sql \\
      > infra/postgres/init/08-taxon-name-zh.sql
"""
from __future__ import annotations

import asyncio
import logging
import sys
from pathlib import Path

import asyncpg
import openpyxl

from app.config import settings

log = logging.getLogger("import_taxon_zh")

XLSX_PATH = Path(__file__).resolve().parent.parent / "data_import" / "fenlei.xlsx"

RANK_MAP = {
    1: "class",
    2: "subclass",
    3: "infraclass",
    4: "superorder",
    5: "order",
    6: "suborder",
    7: "infraorder",
}

SEED_ENTRIES: list[tuple[str, str, str]] = [
    ("Animalia", "动物界", "kingdom"),
    ("Mollusca", "软体动物门", "phylum"),
]


def parse_xlsx() -> list[tuple[str, str, str]]:
    wb = openpyxl.load_workbook(str(XLSX_PATH))
    ws = wb["Sheet2"]

    entries: list[tuple[str, str, str]] = []

    for row in range(4, ws.max_row + 1):
        for col in range(1, 9):
            val = ws.cell(row, col).value
            if not val or not val.strip():
                continue
            val = val.strip()
            parts = val.rsplit(" ", 1)
            if len(parts) != 2:
                continue
            chinese, latin = parts[0].strip(), parts[1].strip()
            if not latin or not latin[0].isupper() or not chinese:
                continue
            rank_type = RANK_MAP.get(col, "other")
            entries.append((latin, chinese, rank_type))

        latin_family = ws.cell(row, 10).value
        chinese_family = ws.cell(row, 12).value
        if latin_family and latin_family.strip() and chinese_family and chinese_family.strip():
            entries.append((latin_family.strip(), chinese_family.strip(), "family"))

    wb.close()
    return entries


def _emit_sql(entries: list[tuple[str, str, str]]) -> None:
    print("-- AUTO-GENERATED from data_import/fenlei.xlsx + SEED_ENTRIES")
    print("-- Regenerate: python -m scripts.import_taxon_zh --emit-sql")
    print()
    print("CREATE TABLE IF NOT EXISTS taxon_name_zh (")
    print("    latin_name   TEXT PRIMARY KEY,")
    print("    chinese_name TEXT NOT NULL,")
    print("    rank_type    TEXT NOT NULL")
    print(");")
    print()
    print("CREATE INDEX IF NOT EXISTS idx_taxon_name_zh_rank ON taxon_name_zh (rank_type);")
    print()
    print("INSERT INTO taxon_name_zh (latin_name, chinese_name, rank_type) VALUES")
    deduped: dict[str, tuple[str, str, str]] = {}
    for entry in entries:
        deduped[entry[0]] = entry
    rows = sorted(deduped.values(), key=lambda e: (e[2], e[0]))
    lines = [f"  ('{l.replace(chr(39), chr(39)*2)}', '{c.replace(chr(39), chr(39)*2)}', '{r}')" for l, c, r in rows]
    print(",\n".join(lines))
    print("ON CONFLICT (latin_name) DO NOTHING;")


async def run() -> int:
    emit_sql_mode = "--emit-sql" in sys.argv
    log_stream = sys.stderr if emit_sql_mode else sys.stdout
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-7s %(message)s", stream=log_stream)

    if not XLSX_PATH.exists():
        log.error("xlsx not found: %s", XLSX_PATH)
        return 1

    entries = parse_xlsx()
    entries.extend(SEED_ENTRIES)
    log.info("parsed %d entries from xlsx", len(entries))

    if emit_sql_mode:
        _emit_sql(entries)
        return 0

    conn = await asyncpg.connect(settings.DATABASE_URL_SYNC)
    try:
        await conn.execute("DELETE FROM taxon_name_zh")
        await conn.executemany(
            """INSERT INTO taxon_name_zh (latin_name, chinese_name, rank_type)
               VALUES ($1, $2, $3)
               ON CONFLICT (latin_name) DO UPDATE
               SET chinese_name = EXCLUDED.chinese_name, rank_type = EXCLUDED.rank_type""",
            entries,
        )
        total = await conn.fetchval("SELECT COUNT(*) FROM taxon_name_zh")
        log.info("imported %d entries into taxon_name_zh", total)
    finally:
        await conn.close()
    return 0


def main() -> int:
    return asyncio.run(run())


if __name__ == "__main__":
    sys.exit(main())
