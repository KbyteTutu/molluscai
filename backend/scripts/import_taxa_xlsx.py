#!/usr/bin/env python3
"""
Import legacy Mollusca_species.xlsx into the `taxa` table.

Each row has a full classification chain + a WoRMS URL containing the
AphiaID. We parse the AphiaID, derive scientificname/authority/genus/
species from the Species string, and bulk-upsert into Postgres.

Row counts: ~141k. Expect ~10–30 seconds end-to-end.

Invoked via: docker compose exec backend python -m scripts.import_taxa_xlsx /tmp/m.xlsx
"""
from __future__ import annotations

import asyncio
import logging
import re
import sys
from pathlib import Path
from typing import Iterator, Optional

import asyncpg
import openpyxl

from app.config import settings

log = logging.getLogger("import_taxa_xlsx")

APHIA_RE = re.compile(r"id=(\d+)")
AUTHORITY_SPLIT = re.compile(r"^(?P<sci>[A-Z][a-zA-Z\-\.\s\u00c0-\u024f]+?)\s+(?P<auth>\(?[A-Z].*)$")


def parse_aphia_id(url: str) -> Optional[int]:
    if not url:
        return None
    m = APHIA_RE.search(url)
    return int(m.group(1)) if m else None


def split_sci_authority(raw: str) -> tuple[str, Optional[str]]:
    if not raw:
        return "", None
    clean = raw.replace("\u00a0", " ").replace("†", "").strip()
    m = AUTHORITY_SPLIT.match(clean)
    if m:
        return m.group("sci").strip(), m.group("auth").strip()
    return clean, None


def extract_genus_species(sciname: str) -> tuple[Optional[str], Optional[str]]:
    parts = sciname.split()
    if len(parts) < 2:
        return (parts[0] if parts else None), None
    return parts[0], parts[1]


def rows_from_xlsx(path: Path) -> Iterator[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        species_raw, url, fossil, genus, family_raw, superfamily_raw, infraorder, suborder, order_, superorder, _subterclass, infraclass, subclass, class_, subphylum, phylum, _node, _family_url = row

        aphia_id = parse_aphia_id(url or "")
        if aphia_id is None:
            continue

        sciname, authority = split_sci_authority(species_raw or "")
        _genus, species_epithet = extract_genus_species(sciname)

        family = (family_raw or "").split(" ", 1)[0] or None
        superfamily = (superfamily_raw or "").split(" ", 1)[0] or None

        yield {
            "aphia_id": aphia_id,
            "scientificname": sciname or None,
            "authority": authority,
            "rank": "Species" if species_epithet else None,
            "status": None,
            "parent_aphia_id": None,
            "kingdom": "Animalia",
            "phylum": phylum,
            "subphylum": subphylum,
            "class_": class_,
            "subclass": subclass,
            "infraclass": infraclass,
            "superorder": superorder,
            "order_": order_,
            "suborder": suborder,
            "infraorder": infraorder,
            "superfamily": superfamily,
            "family": family,
            "genus": genus or _genus,
            "species_epithet": species_epithet,
            "is_extinct": bool(fossil) or None,
            "url": url,
            "data_source": "xlsx",
        }


async def bulk_upsert(conn: asyncpg.Connection, records: list[dict]) -> int:
    if not records:
        return 0
    tuples = [
        (
            r["aphia_id"], r["scientificname"], r["authority"], r["rank"], r["status"],
            r["parent_aphia_id"],
            r["kingdom"], r["phylum"], r["subphylum"], r["class_"], r["subclass"],
            r["infraclass"], r["superorder"], r["order_"], r["suborder"],
            r["infraorder"], r["superfamily"], r["family"], r["genus"],
            r["species_epithet"], r["is_extinct"], r["url"], r["data_source"],
        )
        for r in records
    ]
    await conn.executemany(
        """
        INSERT INTO taxa (aphia_id, scientificname, authority, rank, status,
            parent_aphia_id,
            kingdom, phylum, subphylum, class, subclass,
            infraclass, superorder, "order", suborder,
            infraorder, superfamily, family, genus,
            species_epithet, is_extinct, url, data_source)
        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17,$18,$19,$20,$21,$22,$23)
        ON CONFLICT (aphia_id) DO UPDATE SET
            scientificname = COALESCE(taxa.scientificname, EXCLUDED.scientificname),
            authority = COALESCE(taxa.authority, EXCLUDED.authority),
            phylum = COALESCE(taxa.phylum, EXCLUDED.phylum),
            subphylum = COALESCE(taxa.subphylum, EXCLUDED.subphylum),
            class = COALESCE(taxa.class, EXCLUDED.class),
            subclass = COALESCE(taxa.subclass, EXCLUDED.subclass),
            infraclass = COALESCE(taxa.infraclass, EXCLUDED.infraclass),
            superorder = COALESCE(taxa.superorder, EXCLUDED.superorder),
            "order" = COALESCE(taxa."order", EXCLUDED."order"),
            suborder = COALESCE(taxa.suborder, EXCLUDED.suborder),
            infraorder = COALESCE(taxa.infraorder, EXCLUDED.infraorder),
            superfamily = COALESCE(taxa.superfamily, EXCLUDED.superfamily),
            family = COALESCE(taxa.family, EXCLUDED.family),
            genus = COALESCE(taxa.genus, EXCLUDED.genus),
            species_epithet = COALESCE(taxa.species_epithet, EXCLUDED.species_epithet),
            is_extinct = COALESCE(taxa.is_extinct, EXCLUDED.is_extinct),
            url = COALESCE(taxa.url, EXCLUDED.url),
            data_source = CASE
                WHEN taxa.data_source = 'worms' THEN 'merged'
                ELSE taxa.data_source
            END,
            updated_at = now()
        """,
        tuples,
    )
    return len(tuples)


async def run(xlsx_path: Path) -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-7s %(message)s")
    dsn = settings.DATABASE_URL_SYNC
    conn = await asyncpg.connect(dsn)
    try:
        log.info("reading %s ...", xlsx_path)
        batch: list[dict] = []
        total = 0
        async with conn.transaction():
            for r in rows_from_xlsx(xlsx_path):
                batch.append(r)
                if len(batch) >= 2000:
                    total += await bulk_upsert(conn, batch)
                    batch = []
                    log.info("upserted %d rows ...", total)
            if batch:
                total += await bulk_upsert(conn, batch)
        final = await conn.fetchrow("SELECT COUNT(*) AS n, COUNT(*) FILTER (WHERE data_source='xlsx') AS xlsx_only FROM taxa")
        log.info("done. inserted/updated: %d  taxa total: %d  xlsx-only: %d",
                 total, final["n"], final["xlsx_only"])
    finally:
        await conn.close()


def main() -> int:
    if len(sys.argv) != 2:
        print("usage: python -m scripts.import_taxa_xlsx <path/to/Mollusca_species.xlsx>", file=sys.stderr)
        return 2
    asyncio.run(run(Path(sys.argv[1])))
    return 0


if __name__ == "__main__":
    sys.exit(main())
